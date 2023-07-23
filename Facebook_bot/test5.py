import time
from selenium import webdriver
from bs4 import BeautifulSoup
import pickle
import time
from bs4 import BeautifulSoup
import requests
import openpyxl
import pytest
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import pandas as pd
import time
from selenium import webdriver
from bs4 import BeautifulSoup

from webscrping.scrape import soup

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Members'
print(excel.sheetnames)  # rename sheet
sheet.append(['name'])


def scrape_facebook_group_members(url):
    chrome_options = Options()
    chrome_options.add_argument("--remote-debugging-port=9222")
    chrome_options.add_argument(
        'user-data-dir=C:\\Users\\KloverCloud\\AppData\\Local\\Programs\\Python\\Python37-32\\chromedriver_win32\\chromeprofile')
    chrome_driver = "C:\\Users\\KloverCloud\\AppData\\Local\\Programs\\Python\\Python37-32\\chromedriver_win32\\chromedriver.exe"
    driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)  # driver = webdriver.Chrome()
    driver.maximize_window()
    # driver.get("https://www.facebook.com/groups/261000217971523")

    # Start the Selenium WebDriver (Chrome in this example)
    driver.get(url)

    # Wait for the page to load (5 seconds in this example)
    time.sleep(5)

    # Click on the "Members" tab
    members_tab = driver.find_element_by_link_text('People')
    members_tab.click()

    # Wait for the page to load (5 seconds in this example)
    time.sleep(5)

    # Scroll down to load more members (repeat the scroll 3 times with 3 seconds wait each)
    # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    # time.sleep(5)

    for _ in range(3):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(10)

    # Get the page source after scrolling and pass it to BeautifulSoup
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, 'html.parser')
    time.sleep(2)

    # Assuming member names are stored inside <a> tags with specific classes
    # You'll need to inspect the Facebook page source to identify the appropriate elements
    # and update the code accordingly.
    # Example: if member names are inside <a class="member-link" href="/user123">John Doe</a> tags:

    # member_link_elements = soup.find_all('a', attrs={'role': 'link'})
    # # member_link_elements = soup.find_all('a', class_='member-link')
    # # print(member_link_elements)
    # member_names = [element.text.strip() for element in member_link_elements]

    # Close the WebDriver
    driver.quit()

    return member_names


# Extract group members
group_members = []
member_link_elements = soup.find_all('a', attrs={'role': 'link'})
member_names = [element.text.strip() for element in member_link_elements]

# member_elements = soup.find_all('a', attrs={'role': 'link'})
for member_element in member_names:
    member_name = member_element.text
    if member_name:
        group_members.append(member_name)

# Print the group members
print("Group Members:")
for member in group_members:
    print(member)

# Save group members in an Excel file
df = pd.DataFrame({'Group Members': group_members})
df.to_excel('test.xlsx', index=False)

print("Group members saved in 'group_members.xlsx' Excel file.")
