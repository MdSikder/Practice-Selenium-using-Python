from selenium.webdriver.common.by import By
import openpyxl
from openpyxl import load_workbook
import XLutils
from selenium import webdriver

# driver = webdriver.Chrome("C:\\Users\\KloverCloud\\AppData\\Local\\Programs\\Python\\Python37-32\\chromedriver_win32\\chromedriver.exe")
# driver.get("http://newtours.demoaut.com/")
# driver.maximize_window()

path = "C:\\Users\\KloverCloud\\PycharmProjects\\Practice_Selenium_Using_Python\\Facebook_bot\\excel\\hello.xlsx"
path2 = "C:\\Users\\KloverCloud\\PycharmProjects\\Practice_Selenium_Using_Python\\Facebook_bot\\test.xlsx"
# rows = XLutils.getRowCount(path, "Sheet1")

workbook = openpyxl.load_workbook(path2)

# workbook = load_workbook(path)
sheet = workbook.active

rows1 = sheet.max_row
cols = sheet.max_column
print(rows1)
print(cols)

# for r in range(1, rows1 + 1):
#     for c in range(1, cols + 1):
#         print(sheet.cell(row=r, column=c).value, end="  ")

for r in range(1, rows1 + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(row=r, column=c).value)

# for r in range(2, rows1+1):
# for c in range(1, cols + 15):
#     username = XLutils.readData(path, "sheet1", r, 1)
#     print(username)
#     # XLutils.readData(path, "Sheet1", r, 2)
#     # search_bar = driver.find_element_by_xpath("")
#     # search_bar.click()
#     # search_bar.send_keys(username)
